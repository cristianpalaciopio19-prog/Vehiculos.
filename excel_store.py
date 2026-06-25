import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'data', 'viajes.xlsx')

COLUMNAS = [
    'ID', 'Fecha registro', 'Placa', 'Número interno', 'Ruta', 'Municipio',
    'Fecha viaje', 'Turno', 'Semana', 'Conductor', 'Kilómetros',
    'Ingreso/Salida', 'Quién diligencia'
]

CAMPOS = [
    'id', 'fecha_registro', 'placa', 'numero_interno', 'ruta', 'municipio',
    'fecha_viaje', 'turno', 'semana', 'conductor', 'km',
    'lleva_trae_personal', 'quien_diligencia'
]

HEADER_FILL = PatternFill('solid', start_color='3B6D11')
HEADER_FONT = Font(color='FFFFFF', bold=True, name='Arial', size=11)
CELL_FONT = Font(name='Arial', size=10)


def inicializar_excel():
    if os.path.exists(EXCEL_PATH):
        return
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Viajes'
    sheet.append(COLUMNAS)
    for col_idx in range(1, len(COLUMNAS) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
    anchos = [6, 16, 10, 12, 10, 14, 12, 8, 10, 20, 11, 16, 18]
    for i, ancho in enumerate(anchos, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = ancho
    sheet.freeze_panes = 'A2'
    wb.save(EXCEL_PATH)


def siguiente_id(sheet):
    max_id = 0
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and isinstance(row[0], int):
            max_id = max(max_id, row[0])
    return max_id + 1


def agregar_viaje(placa, numero_interno, ruta, municipio, fecha_viaje, turno, semana,
                   conductor, km, lleva_trae_personal, quien_diligencia):
    inicializar_excel()
    wb = load_workbook(EXCEL_PATH)
    sheet = wb['Viajes']
    nuevo_id = siguiente_id(sheet)
    fila = [
        nuevo_id,
        datetime.now().strftime('%d/%m/%Y %H:%M'),
        placa, numero_interno, ruta, municipio, fecha_viaje, turno, semana,
        conductor, km, lleva_trae_personal, quien_diligencia
    ]
    sheet.append(fila)
    row_idx = sheet.max_row
    for col_idx in range(1, len(COLUMNAS) + 1):
        sheet.cell(row=row_idx, column=col_idx).font = CELL_FONT
    wb.save(EXCEL_PATH)
    return nuevo_id


def listar_viajes(filtro_placa=None, filtro_municipio=None):
    inicializar_excel()
    wb = load_workbook(EXCEL_PATH)
    sheet = wb['Viajes']
    viajes = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        viaje = dict(zip(CAMPOS, row))
        if filtro_placa and filtro_placa.upper() not in (viaje['placa'] or '').upper():
            continue
        if filtro_municipio and filtro_municipio != viaje['municipio']:
            continue
        viajes.append(viaje)
    viajes.sort(key=lambda v: v['id'], reverse=True)
    return viajes


def detalle_placa(placa):
    todos = listar_viajes()
    return [v for v in todos if (v['placa'] or '').upper() == placa.upper()]


def eliminar_viaje(viaje_id):
    inicializar_excel()
    wb = load_workbook(EXCEL_PATH)
    sheet = wb['Viajes']
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == viaje_id:
            sheet.delete_rows(row[0].row, 1)
            break
    wb.save(EXCEL_PATH)


def resumen():
    viajes = listar_viajes()
    placas = set(v['placa'] for v in viajes if v['placa'])
    total_km = sum(v['km'] or 0 for v in viajes)
    return {
        'total_viajes': len(viajes),
        'placas_distintas': len(placas),
        'total_km': round(total_km, 1),
    }
